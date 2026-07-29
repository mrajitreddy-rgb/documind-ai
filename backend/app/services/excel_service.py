from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


EXPORT_FOLDER = Path("exports/excel")
EXPORT_FOLDER.mkdir(parents=True, exist_ok=True)


def create_excel(invoices):
    """
    Create an Excel file from extracted invoices.

    Returns:
        Path to the generated Excel file.
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"
    worksheet.freeze_panes = "A2"

    headers = [
        "Invoice Number",
        "Invoice Date",
        "Dealer",
        "Customer",
        "Amount",
        "Page"
    ]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="00B0F0"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Write headers
    for column, header in enumerate(headers, start=1):

        cell = worksheet.cell(
            row=1,
            column=column
        )

        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    worksheet.row_dimensions[1].height = 24
    # Write invoice rows
    for row, invoice in enumerate(invoices, start=2):

        worksheet.cell(row=row, column=1).value = invoice.get("invoice_number", "")
        worksheet.cell(row=row, column=2).value = invoice.get("invoice_date", "")
        worksheet.cell(row=row, column=3).value = invoice.get("dealer", "")
        worksheet.cell(row=row, column=4).value = invoice.get("customer", "")
        worksheet.cell(row=row, column=5).value = invoice.get("amount", "")
        worksheet.cell(row=row, column=6).value = invoice.get("page", "")
        worksheet.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        worksheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

    # Auto-size columns
    for column_cells in worksheet.columns:

        length = 0
        column = column_cells[0].column

        for cell in column_cells:

            try:
                length = max(length, len(str(cell.value)))
            except Exception:
                pass

        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = length + 4

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f"DocuMind_Invoices_{timestamp}.xlsx"

    filepath = EXPORT_FOLDER / filename
    worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(filepath)

    return filepath